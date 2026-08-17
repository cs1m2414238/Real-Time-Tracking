"""
Synthetic Military Logistics & Readiness Data Generator
Generates realistic, unclassified synthetic dataset for BI portfolio demonstrations across
Excel, SQL, Power BI, Tableau, and Python.
"""

import os
import random
import datetime
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Set fixed seed for deterministic reproducibility
random.seed(42)
np.random.seed(42)

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
os.makedirs(DATA_DIR, exist_ok=True)

# 1. Reference Data: Bases
BASES = [
    {"Base_ID": "B01", "Base_Name": "Lucknow", "Region": "Central Command", "Strategic_Role": "Logistics & Reserve Hub", "Commander_Rank": "Maj General", "Target_Readiness": 90.0},
    {"Base_ID": "B02", "Base_Name": "Kanpur", "Region": "Central Command", "Strategic_Role": "Heavy Maintenance Depot", "Commander_Rank": "Brigadier", "Target_Readiness": 85.0},
    {"Base_ID": "B03", "Base_Name": "Bhopal", "Region": "Central Command", "Strategic_Role": "Armor & Training Center", "Commander_Rank": "Maj General", "Target_Readiness": 85.0},
    {"Base_ID": "B04", "Base_Name": "Delhi", "Region": "Northern Command", "Strategic_Role": "Strategic Air Defense Hub", "Commander_Rank": "Lt General", "Target_Readiness": 92.0},
    {"Base_ID": "B05", "Base_Name": "Leh", "Region": "High-Altitude Ops", "Strategic_Role": "Forward Mountain Command", "Commander_Rank": "Maj General", "Target_Readiness": 88.0},
    {"Base_ID": "B06", "Base_Name": "Jodhpur", "Region": "Western Command", "Strategic_Role": "Desert Warfare Logistics", "Commander_Rank": "Brigadier", "Target_Readiness": 87.0},
    {"Base_ID": "B07", "Base_Name": "Siliguri", "Region": "Eastern Command", "Strategic_Role": "Corridor Transit Logistics", "Commander_Rank": "Brigadier", "Target_Readiness": 86.0},
    {"Base_ID": "B08", "Base_Name": "Pune", "Region": "Southern Command", "Strategic_Role": "Technical Training & Reserves", "Commander_Rank": "Maj General", "Target_Readiness": 89.0}
]

# 2. Reference Data: Equipment Types
EQUIPMENT_TYPES = [
    {"Type_ID": "ET01", "Equipment_Type": "Heavy Tactical Truck", "Category": "Transport Fleet", "Fuel_Burn_Rate_L_Hr": 2.6, "Service_Interval_Days": 60, "Base_Cost_INR": 12500},
    {"Type_ID": "ET02", "Equipment_Type": "Armored Personnel Carrier (APC)", "Category": "Combat Armor", "Fuel_Burn_Rate_L_Hr": 4.8, "Service_Interval_Days": 90, "Base_Cost_INR": 48000},
    {"Type_ID": "ET03", "Equipment_Type": "Field Diesel Generator", "Category": "Power Logistics", "Fuel_Burn_Rate_L_Hr": 1.4, "Service_Interval_Days": 45, "Base_Cost_INR": 24000},
    {"Type_ID": "ET04", "Equipment_Type": "Tactical UAV Launcher", "Category": "Surveillance Systems", "Fuel_Burn_Rate_L_Hr": 1.1, "Service_Interval_Days": 30, "Base_Cost_INR": 62000},
    {"Type_ID": "ET05", "Equipment_Type": "Mobile Radar Unit", "Category": "Air Defense", "Fuel_Burn_Rate_L_Hr": 2.2, "Service_Interval_Days": 60, "Base_Cost_INR": 85000},
    {"Type_ID": "ET06", "Equipment_Type": "Artillery Tractor", "Category": "Heavy Towing", "Fuel_Burn_Rate_L_Hr": 3.5, "Service_Interval_Days": 75, "Base_Cost_INR": 34000}
]

# Weights for realistic distribution
BASE_WEIGHTS = [0.18, 0.17, 0.14, 0.16, 0.11, 0.09, 0.08, 0.07]
EQUIP_TYPE_WEIGHTS = [0.32, 0.18, 0.22, 0.08, 0.07, 0.13]

CURRENT_REF_DATE = datetime.date(2026, 8, 15)

def generate_datasets(num_records=1250):
    print(f"Generating {num_records} synthetic equipment inventory records...")
    
    clean_records = []
    dirty_records = []
    maintenance_records = []
    
    maint_log_counter = 5001

    for i in range(1, num_records + 1):
        eq_id = f"EQ-{1000 + i}"
        
        # Select base and equipment type based on weights
        base_obj = random.choices(BASES, weights=BASE_WEIGHTS, k=1)[0]
        base_name = base_obj["Base_Name"]
        base_id = base_obj["Base_ID"]
        
        eq_type_obj = random.choices(EQUIPMENT_TYPES, weights=EQUIP_TYPE_WEIGHTS, k=1)[0]
        eq_type = eq_type_obj["Equipment_Type"]
        type_id = eq_type_obj["Type_ID"]
        
        # Base-specific nuances
        if base_name == "Kanpur":
            status_pool = ["Operational", "Maintenance", "Out of Service"]
            status_weights = [0.74, 0.19, 0.07]
        elif base_name == "Delhi":
            status_pool = ["Operational", "Maintenance", "Out of Service"]
            status_weights = [0.93, 0.05, 0.02]
        elif base_name == "Lucknow":
            status_pool = ["Operational", "Maintenance", "Out of Service"]
            status_weights = [0.91, 0.07, 0.02]
        elif base_name == "Bhopal":
            status_pool = ["Operational", "Maintenance", "Out of Service"]
            status_weights = [0.78, 0.15, 0.07]
        else:
            status_pool = ["Operational", "Maintenance", "Out of Service"]
            status_weights = [0.85, 0.11, 0.04]
            
        status = random.choices(status_pool, weights=status_weights, k=1)[0]
        
        # Spare parts status
        if base_name == "Bhopal" and random.random() < 0.35:
            spare_parts = random.choices(["Low", "Critical Shortage"], weights=[0.65, 0.35], k=1)[0]
        elif status in ["Maintenance", "Out of Service"] and random.random() < 0.40:
            spare_parts = random.choices(["Low", "Critical Shortage"], weights=[0.60, 0.40], k=1)[0]
        else:
            spare_parts = random.choices(["Available", "Low", "Critical Shortage"], weights=[0.78, 0.16, 0.06], k=1)[0]
            
        # Operating hours
        if eq_type == "Field Diesel Generator":
            operating_hours = int(np.random.normal(loc=260, scale=60))
        elif eq_type == "Heavy Tactical Truck":
            operating_hours = int(np.random.normal(loc=175, scale=45))
        else:
            operating_hours = int(np.random.normal(loc=195, scale=50))
        operating_hours = max(45, min(680, operating_hours))
        
        # Fuel used
        burn_rate = eq_type_obj["Fuel_Burn_Rate_L_Hr"] * random.uniform(0.92, 1.12)
        fuel_used = int(operating_hours * burn_rate)
        
        # Service Dates
        interval = eq_type_obj["Service_Interval_Days"]
        days_since_last_service = random.randint(10, interval + (35 if base_name == "Kanpur" and status != "Operational" else 15))
        last_service_date = CURRENT_REF_DATE - datetime.timedelta(days=days_since_last_service)
        next_service_date = last_service_date + datetime.timedelta(days=interval)
        
        # Maintenance Cost
        base_cost = eq_type_obj["Base_Cost_INR"]
        cost_multiplier = 1.0 + (operating_hours / 350.0)
        
        if eq_type == "Field Diesel Generator":
            cost_multiplier *= random.uniform(1.20, 1.45)
            
        if status == "Maintenance":
            cost_multiplier *= random.uniform(1.35, 1.90)
        elif status == "Out of Service":
            cost_multiplier *= random.uniform(1.70, 2.50)
            
        maintenance_cost = int(base_cost * cost_multiplier * random.uniform(0.90, 1.15))
        maintenance_cost = round(maintenance_cost, -2) # round to nearest 100
        
        # Readiness %
        if status == "Operational":
            readiness_pct = round(random.uniform(80.0, 99.0), 1)
        elif status == "Maintenance":
            readiness_pct = round(random.uniform(45.0, 68.0), 1)
        else: # Out of Service
            readiness_pct = round(random.uniform(15.0, 44.0), 1)
            
        # Downtime Days
        if status == "Operational":
            downtime_days = 0
        elif status == "Maintenance":
            downtime_days = random.randint(3, 21)
        else:
            downtime_days = random.randint(18, 48)
            
        is_critical = 1 if readiness_pct < 60.0 else 0
        
        # Append to Clean Record
        clean_records.append({
            "Equipment_ID": eq_id,
            "Base_ID": base_id,
            "Base": base_name,
            "Type_ID": type_id,
            "Equipment_Type": eq_type,
            "Category": eq_type_obj["Category"],
            "Status": status,
            "Last_Service": last_service_date.strftime("%Y-%m-%d"),
            "Next_Service": next_service_date.strftime("%Y-%m-%d"),
            "Fuel_Used_L": fuel_used,
            "Operating_Hours": operating_hours,
            "Maintenance_Cost_INR": maintenance_cost,
            "Spare_Parts": spare_parts,
            "Readiness_Pct": readiness_pct,
            "Downtime_Days": downtime_days,
            "Is_Critical": is_critical
        })
        
        # Build dirty record version
        base_dirty = base_name
        if random.random() < 0.25:
            casing_style = random.choice(["upper", "lower", "whitespace", "mixed"])
            if casing_style == "upper":
                base_dirty = base_name.upper()
            elif casing_style == "lower":
                base_dirty = base_name.lower()
            elif casing_style == "whitespace":
                base_dirty = f"  {base_name}   "
            elif casing_style == "mixed":
                base_dirty = base_name[0].lower() + base_name[1:].upper()
                
        eq_type_dirty = eq_type
        if random.random() < 0.20:
            eq_type_dirty = random.choice([
                eq_type.upper(),
                eq_type.lower(),
                f" {eq_type} ",
                eq_type.replace(" ", "_")
            ])
            
        cost_dirty = maintenance_cost
        if random.random() < 0.35:
            cost_dirty = random.choice([
                f"₹{maintenance_cost:,}",
                f"{maintenance_cost:,} INR",
                f"Rs. {maintenance_cost}",
                f"{maintenance_cost}"
            ])
            
        readiness_dirty = readiness_pct
        if random.random() < 0.30:
            readiness_dirty = random.choice([
                f"{readiness_pct}%",
                f"{readiness_pct/100.0:.4f}",
                str(readiness_pct),
                f"{int(readiness_pct)}%"
            ])
            
        last_service_dirty = last_service_date.strftime("%Y-%m-%d")
        next_service_dirty = next_service_date.strftime("%Y-%m-%d")
        if random.random() < 0.25:
            fmt = random.choice(["%d-%b-%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y"])
            last_service_dirty = last_service_date.strftime(fmt)
            next_service_dirty = next_service_date.strftime(fmt)
            
        fuel_dirty = fuel_used
        spares_dirty = spare_parts
        if random.random() < 0.04:
            fuel_dirty = None
        if random.random() < 0.03:
            spares_dirty = None

        dirty_records.append({
            "Equipment_ID": eq_id,
            "Base": base_dirty,
            "Equipment_Type": eq_type_dirty,
            "Status": status.lower() if random.random() < 0.15 else status,
            "Last_Service": last_service_dirty,
            "Next_Service": next_service_dirty,
            "Fuel_Used": fuel_dirty,
            "Operating_Hours": operating_hours,
            "Maintenance_Cost": cost_dirty,
            "Spare_Parts": spares_dirty,
            "Readiness": readiness_dirty
        })
        
        # Historical Maintenance Log Records
        num_past_services = random.randint(2, 4)
        for s in range(num_past_services):
            service_offset_days = (s + 1) * interval + random.randint(-8, 12)
            hist_service_date = CURRENT_REF_DATE - datetime.timedelta(days=service_offset_days)
            hist_cost = int(base_cost * random.uniform(0.85, 1.40))
            hist_hours_spent = random.choice([4, 6, 8, 12, 16, 24, 48])
            hist_parts = random.choice([
                "Oil Filter, Gaskets", "Brake Pads, Hydraulic Line", "Injector Nozzle, Fuel Filter",
                "Starter Motor, Battery", "Alternator, Belt Drive", "Radiator Core, Coolant",
                "Sensors, Wiring Harness", "Tire Assembly, Hub Bearing"
            ])
            service_type = random.choice(["Routine Preventive", "Corrective Repair", "Oil & Filter Overhaul", "Emergency Repair", "Pre-Mission Inspection"])
            
            maintenance_records.append({
                "Log_ID": f"ML-{maint_log_counter}",
                "Equipment_ID": eq_id,
                "Base_ID": base_id,
                "Service_Date": hist_service_date.strftime("%Y-%m-%d"),
                "Service_Type": service_type,
                "Technician_Rank": random.choice(["Havildar Tech", "Naik Mech", "Subedar Tech", "Warrant Officer", "Sergeant"]),
                "Labor_Hours": hist_hours_spent,
                "Parts_Replaced": hist_parts,
                "Cost_INR": hist_cost,
                "Downtime_Days": random.choice([1, 2, 3, 5, 7]) if "Repair" in service_type else 1
            })
            maint_log_counter += 1

    # Inject 20 Duplicate rows in dirty dataset
    for _ in range(20):
        dup_idx = random.randint(0, len(dirty_records) - 1)
        dirty_records.append(dirty_records[dup_idx].copy())
        
    random.shuffle(dirty_records)
    
    # Convert to DataFrames
    df_clean = pd.DataFrame(clean_records)
    df_dirty = pd.DataFrame(dirty_records)
    df_bases = pd.DataFrame(BASES)
    df_types = pd.DataFrame(EQUIPMENT_TYPES)
    df_logs = pd.DataFrame(maintenance_records)
    
    # Save CSVs
    df_clean.to_csv(os.path.join(DATA_DIR, "fact_equipment_inventory.csv"), index=False)
    df_dirty.to_csv(os.path.join(DATA_DIR, "raw_equipment_dirty.csv"), index=False)
    df_bases.to_csv(os.path.join(DATA_DIR, "dim_bases.csv"), index=False)
    df_types.to_csv(os.path.join(DATA_DIR, "dim_equipment_types.csv"), index=False)
    df_logs.to_csv(os.path.join(DATA_DIR, "fact_maintenance_logs.csv"), index=False)
    
    print(f"CSVs saved in {DATA_DIR}:")
    print(f" - Clean Inventory: {len(df_clean)} rows")
    print(f" - Dirty Raw Data: {len(df_dirty)} rows")
    print(f" - Bases Master: {len(df_bases)} rows")
    print(f" - Equipment Types Master: {len(df_types)} rows")
    print(f" - Historical Maintenance Logs: {len(df_logs)} rows")
    
    # Generate Excel Workbook
    generate_excel_model(df_clean, df_dirty, df_bases, df_types, df_logs)
    
    return df_clean, df_dirty, df_bases, df_types, df_logs

def generate_excel_model(df_clean, df_dirty, df_bases, df_types, df_logs):
    excel_path = os.path.join(DATA_DIR, "military_logistics_model.xlsx")
    print(f"Creating formatted Excel model at {excel_path}...")
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    regular_font = Font(name="Segoe UI", size=10)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 1. Sheet: Executive_KPI_Summary
    ws_kpi = wb.create_sheet(title="Executive_KPI_Summary")
    ws_kpi.views.sheetView[0].showGridLines = True
    
    ws_kpi.merge_cells("A1:H1")
    title_cell = ws_kpi["A1"]
    title_cell.value = "MILITARY LOGISTICS & READINESS EXECUTIVE BI SUMMARY"
    title_cell.font = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_kpi.row_dimensions[1].height = 40
    
    kpis = [
        ("Total Equipment Fleet", "=COUNTA(Cleaned_Inventory!A2:A1251)", "A3:B4", "Units"),
        ("Operational Assets", '=COUNTIF(Cleaned_Inventory!G2:G1251, "Operational")', "C3:D4", "Units"),
        ("Readiness Rate %", '=C4/A4', "E3:F4", "Percent"),
        ("Under Maintenance", '=COUNTIF(Cleaned_Inventory!G2:G1251, "Maintenance")', "A6:B7", "Units"),
        ("Critical Assets (<60%)", '=COUNTIF(Cleaned_Inventory!P2:P1251, 1)', "C6:D7", "Units"),
        ("Total Maintenance Cost", '=SUM(Cleaned_Inventory!L2:L1251)', "E6:F7", "Currency")
    ]
    
    for label, formula, cell_range, fmt in kpis:
        top_left = cell_range.split(":")[0]
        bot_right = cell_range.split(":")[1]
        col_start = top_left[0]
        row_start = int(top_left[1:])
        col_end = bot_right[0]
        row_end = int(bot_right[1:])
        
        ws_kpi.merge_cells(f"{col_start}{row_start}:{col_end}{row_start}")
        lbl_c = ws_kpi[f"{col_start}{row_start}"]
        lbl_c.value = label
        lbl_c.font = Font(name="Segoe UI", size=9, bold=True, color="64748B")
        lbl_c.alignment = Alignment(horizontal="center", vertical="center")
        
        ws_kpi.merge_cells(f"{col_start}{row_end}:{col_end}{row_end}")
        val_c = ws_kpi[f"{col_start}{row_end}"]
        val_c.value = formula
        val_c.font = Font(name="Segoe UI", size=14, bold=True, color="0F172A")
        val_c.alignment = Alignment(horizontal="center", vertical="center")
        val_c.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        
        if fmt == "Percent":
            val_c.number_format = "0.0%"
        elif fmt == "Currency":
            val_c.number_format = "₹#,##0"
        elif fmt == "Units":
            val_c.number_format = "#,##0"

    ws_kpi["A10"] = "Base Readiness & Maintenance Matrix (Excel Formulas: COUNTIFS, SUMIFS, AVERAGEIFS)"
    ws_kpi["A10"].font = Font(name="Segoe UI", size=12, bold=True, color="0F172A")
    
    headers_base_table = ["Base ID", "Base Name", "Region", "Total Assets", "Operational", "Readiness %", "Total Maint Cost (INR)", "Avg Downtime (Days)"]
    for col_idx, h in enumerate(headers_base_table, start=1):
        cell = ws_kpi.cell(row=11, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    for row_idx, base in enumerate(BASES, start=12):
        b_name = base["Base_Name"]
        b_id = base["Base_ID"]
        reg = base["Region"]
        
        ws_kpi.cell(row=row_idx, column=1, value=b_id).alignment = Alignment(horizontal="center")
        ws_kpi.cell(row=row_idx, column=2, value=b_name).alignment = Alignment(horizontal="left")
        ws_kpi.cell(row=row_idx, column=3, value=reg).alignment = Alignment(horizontal="left")
        
        c_tot = ws_kpi.cell(row=row_idx, column=4, value=f'=COUNTIF(Cleaned_Inventory!C2:C1251, "{b_name}")')
        c_tot.number_format = "#,##0"
        
        c_ops = ws_kpi.cell(row=row_idx, column=5, value=f'=COUNTIFS(Cleaned_Inventory!C2:C1251, "{b_name}", Cleaned_Inventory!G2:G1251, "Operational")')
        c_ops.number_format = "#,##0"
        
        c_rd = ws_kpi.cell(row=row_idx, column=6, value=f'=E{row_idx}/D{row_idx}')
        c_rd.number_format = "0.0%"
        
        c_cost = ws_kpi.cell(row=row_idx, column=7, value=f'=SUMIF(Cleaned_Inventory!C2:C1251, "{b_name}", Cleaned_Inventory!L2:L1251)')
        c_cost.number_format = "₹#,##0"
        
        c_dt = ws_kpi.cell(row=row_idx, column=8, value=f'=AVERAGEIF(Cleaned_Inventory!C2:C1251, "{b_name}", Cleaned_Inventory!O2:O1251)')
        c_dt.number_format = "0.0"
        
        for c in range(1, 9):
            ws_kpi.cell(row=row_idx, column=c).border = thin_border
            ws_kpi.cell(row=row_idx, column=c).font = regular_font

    # 2. Sheet: Cleaned_Inventory
    ws_clean = wb.create_sheet(title="Cleaned_Inventory")
    ws_clean.views.sheetView[0].showGridLines = True
    
    headers_clean = list(df_clean.columns)
    for col_idx, col_name in enumerate(headers_clean, start=1):
        cell = ws_clean.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    for row_idx, row_data in enumerate(df_clean.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_clean.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if headers_clean[col_idx-1] == "Maintenance_Cost_INR":
                cell.number_format = "₹#,##0"
            elif headers_clean[col_idx-1] == "Readiness_Pct":
                cell.number_format = "0.0"
            elif headers_clean[col_idx-1] in ["Fuel_Used_L", "Operating_Hours", "Downtime_Days"]:
                cell.number_format = "#,##0"

    # 3. Sheet: Formulas_Showcase
    ws_formulas = wb.create_sheet(title="Formulas_Showcase")
    ws_formulas.views.sheetView[0].showGridLines = True
    
    ws_formulas["A1"] = "EXCEL ADVANCED FORMULAS DEMONSTRATION & LOOKUP BENCHMARK"
    ws_formulas["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="1E293B")
    
    ws_formulas["A3"] = "1. Single Asset Lookup Engine (Demonstrating XLOOKUP & INDEX-MATCH)"
    ws_formulas["A3"].font = bold_font
    
    ws_formulas["A5"] = "Input Equipment ID:"
    ws_formulas["B5"] = "EQ-1003"
    ws_formulas["B5"].font = Font(name="Segoe UI", size=11, bold=True, color="0284C7")
    ws_formulas["B5"].fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    
    lookup_table = [
        ("Base Name (XLOOKUP)", '=XLOOKUP(B5, Cleaned_Inventory!A2:A1251, Cleaned_Inventory!C2:C1251, "Not Found")'),
        ("Equipment Type (XLOOKUP)", '=XLOOKUP(B5, Cleaned_Inventory!A2:A1251, Cleaned_Inventory!E2:E1251, "Not Found")'),
        ("Operational Status (INDEX-MATCH)", '=INDEX(Cleaned_Inventory!G2:G1251, MATCH(B5, Cleaned_Inventory!A2:A1251, 0))'),
        ("Readiness % (INDEX-MATCH)", '=INDEX(Cleaned_Inventory!N2:N1251, MATCH(B5, Cleaned_Inventory!A2:A1251, 0))'),
        ("Maintenance Cost (XLOOKUP)", '=XLOOKUP(B5, Cleaned_Inventory!A2:A1251, Cleaned_Inventory!L2:L1251, 0)'),
        ("Operating Hours (XLOOKUP)", '=XLOOKUP(B5, Cleaned_Inventory!A2:A1251, Cleaned_Inventory!K2:K1251, 0)'),
        ("Spare Parts Status (INDEX-MATCH)", '=INDEX(Cleaned_Inventory!M2:M1251, MATCH(B5, Cleaned_Inventory!A2:A1251, 0))')
    ]
    
    for idx, (prop, formula) in enumerate(lookup_table, start=6):
        ws_formulas.cell(row=idx, column=1, value=prop).font = bold_font
        c = ws_formulas.cell(row=idx, column=2, value=formula)
        c.font = regular_font
        if "Cost" in prop:
            c.number_format = "₹#,##0"
        elif "Readiness" in prop:
            c.number_format = "0.0"

    ws_formulas["A15"] = "2. Modern Dynamic Array Formulas (Excel 365 / 2021+)"
    ws_formulas["A15"].font = bold_font
    
    dyn_formulas = [
        ("Get Unique Base List", "=UNIQUE(Cleaned_Inventory!C2:C1251)"),
        ("Get Sorted Unique Equipment Types", "=SORT(UNIQUE(Cleaned_Inventory!E2:E1251))"),
        ("Filter All Critical Assets (<60% Readiness)", '=FILTER(Cleaned_Inventory!A2:N1251, Cleaned_Inventory!P2:P1251=1, "None")'),
        ("Filter Kanpur Vehicles Under Maintenance", '=FILTER(Cleaned_Inventory!A2:L1251, (Cleaned_Inventory!C2:C1251="Kanpur")*(Cleaned_Inventory!G2:G1251="Maintenance"), "None")')
    ]
    
    for idx, (desc, formula_text) in enumerate(dyn_formulas, start=17):
        ws_formulas.cell(row=idx, column=1, value=desc).font = bold_font
        ws_formulas.cell(row=idx, column=2, value=formula_text).font = Font(name="Consolas", size=10, color="0369A1")

    # 4. Sheet: Raw_Dirty_Logistics
    ws_dirty = wb.create_sheet(title="Raw_Dirty_Logistics")
    ws_dirty.views.sheetView[0].showGridLines = True
    headers_dirty = list(df_dirty.columns)
    for col_idx, col_name in enumerate(headers_dirty, start=1):
        cell = ws_dirty.cell(row=1, column=col_idx, value=col_name)
        cell.fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for row_idx, row_data in enumerate(df_dirty.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_dirty.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font

    # 5. Sheet: Dim_Bases
    ws_b = wb.create_sheet(title="Dim_Bases")
    for col_idx, col_name in enumerate(df_bases.columns, start=1):
        cell = ws_b.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, row_data in enumerate(df_bases.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row_data, start=1):
            ws_b.cell(row=row_idx, column=col_idx, value=val).font = regular_font

    # 6. Sheet: Dim_Equipment_Types
    ws_t = wb.create_sheet(title="Dim_Equipment_Types")
    for col_idx, col_name in enumerate(df_types.columns, start=1):
        cell = ws_t.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, row_data in enumerate(df_types.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row_data, start=1):
            ws_t.cell(row=row_idx, column=col_idx, value=val).font = regular_font

    # Auto-adjust column widths across all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len and len(val_str) < 50:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(excel_path)
    print(f"Successfully generated {excel_path} with 6 comprehensive worksheets.")

if __name__ == "__main__":
    generate_datasets(1250)
